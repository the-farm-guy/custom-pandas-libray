import openpyxl
from openpyxl import Workbook
import os

class ExcelInputOutput():

    @classmethod
    def read_excel(cls, file_path):
        if not os.path.exists(file_path):
            raise FileNotFoundError('could not find file')
        
        workbook = openpyxl.load_workbook(file_path)
        work_sheet = workbook.active

        columns = [column.value for column in work_sheet[1]]
        data = []
        row_dict = {}

        for rows in work_sheet.iter_rows(values_only = True, min_row = 2):
            if any(cell is not None for cell in rows):
                row_dict = {columns[i] : column for i, column in enumerate(rows)}
                data.append(row_dict)
        return data
    
    @classmethod
    def to_excel(cls, file_path, data ):
        if not data:
            raise ValueError('no data found')
        
        work_book = Workbook()
        work_session = work_book.active

        columns = list(data[0].keys())
        work_session.append(columns)

        for row in data:
            work_session.append([row[col] for col in columns])
    
        work_book.save(file_path)
        print(f'data has been written to the : {file_path}')