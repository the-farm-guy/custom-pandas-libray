import csv
import os
import json 
from  openpyxl import Workbook
import openpyxl

class ReadCsvJson:
    def __init__(self, file_path = None, dtype = None, delimiter = ',', skiprows = 0, header = True):
        self.file_path = file_path
        self.dtype = dtype or {}
        self.delimiter = delimiter
        self.skiprows = skiprows
        self.header = header
        self.data = []
        
        if file_path:
            if file_path.endswith('.csv'):
                self.data = self.read_csv()
            elif file_path.endswith('.json'):
                self.data = self.read_json(file_path = file_path)
            elif file_path.endswith('.xlsx'):
                self.data = self.read_excel(file_path = file_path)

    def __str__(self):
        """Return a string representation when print() is called on the object"""
        return self.as_table()
    
    def __repr__(self):
        """Return a string representation for debugging"""
        return f"CSVProcessor(file_path='{self.file_path}', rows={len(self.data)})"
    
    def __len__(self):
        return len(self.data)
    
    def __getitem__(self, key):
        """Enable slicing with [start:stop] syntax, single item access with [index], or column access with ['column_name']"""
        
        if isinstance(key, slice):    
            result = ReadCsvJson()
            result.data = self.data[key]
            return result
        
        elif isinstance(key, str):
            if not self.data:
                return []
            if key not in self.columns():
                raise KeyError(f"Column '{key}' not found. Available columns: {', '.join(self.columns())}")
            
            result_single = ReadCsvJson()
            result_single.data = [{key : row[key]} for row in self.data]
            return result_single
        
        elif isinstance(key, list):  
            if not all(isinstance(col, str) for col in key): 
                raise TypeError("All column names must be strings.")
            
            missing_columns = [col for col in key if col not in self.columns()]
            if missing_columns:
                raise KeyError(f"Columns {missing_columns} not found. Available columns: {', '.join(self.columns())}")

            result = ReadCsvJson()
            result.data = [{col: row[col] for col in key} for row in self.data]
            return result
        
        elif isinstance(key, int):  
            return self.data[key]
        
        else:
            return self.data[key]
    
    def read_csv(self, file_path=None, dtype=None, delimiter=None, skiprows=None, header=None):
        """Read a CSV file and return a list of dictionaries."""
        
        file_path = file_path or self.file_path
        dtype = dtype or self.dtype
        delimiter = delimiter or self.delimiter
        skiprows = skiprows if skiprows is not None else self.skiprows
        header = header if header is not None else self.header
        
        if not file_path:
            raise ValueError("No file path provided")
            
        if os.path.exists(file_path):
            with open(file_path, 'r') as file:
                reader = csv.reader(file, delimiter=delimiter)
                rows = list(reader)

        else:
            raise FileNotFoundError("couldn't find the file")

        if skiprows > 0:
            rows = rows[skiprows:]

        if header:
            if len(rows) > 0:
                columns = rows[0]
                rows = rows[1:]
            else:
                return []
        else:
            if len(rows) > 0:
                columns = [f"column_{i}" for i in range(len(rows[0]))]
            else:
                return []

        data = []
        for row in rows:
            raw_dict = {}
            for i, col in enumerate(columns):
                value = row[i]
                if dtype and col in dtype:
                    try:
                        value = dtype[col](value)
                    except:
                        raise ValueError(f"Failed to convert column '{col}' to {dtype[col].__name__} for value '{value}'")
                raw_dict[col] = value
            data.append(raw_dict)
        
        self.data = data
        # print(self.data)
        return data
    
    def as_table(self, data=None, columns=None):
        """Format data as a string table."""

        data = data or self.data
        if not data:
            return "Empty dataset"
            
        if columns is None:
            columns = data[0].keys()

        col_widths = {col: max(len(col), max(len(str(row.get(col, ''))) for row in data)) for col in columns}
        header = " | ".join(col.ljust(col_widths[col]) for col in columns)
        separator = "-+-".join('-' * col_widths[col] for col in columns)
        
        rows = []
        for row in data:
            row_str = " | ".join(str(row.get(col, '')).ljust(col_widths[col]) for col in columns)
            rows.append(row_str)

        table = f"{header}\n{separator}\n" + "\n".join(rows)
        return table
    
    def head(self, n=5):
        """Return the first n rows of data as a formatted table."""

        return self.as_table(self.data[:n])
    
    def tail(self, n=5):
        """Return the last n rows of data as a formatted table."""

        return self.as_table(self.data[-n:])
    
    def columns(self):
        """Return the column names of the data."""

        if not self.data:
            return []
        return list(self.data[0].keys())
    
    def shape(self):
        """Return the shape of the data as (rows, columns)."""

        if not self.data:
            return (0, 0)
        return (len(self.data), len(list(self.data[0].keys())))
    
    def to_csv(self, file_path, delimiter = ','):
        if not self.data:
            raise ValueError('no data found')
        
        with open(file_path, 'w') as file:
            fieldnames = self.columns()
            writer = csv.DictWriter(file, fieldnames = fieldnames, delimiter = delimiter)
            writer.writeheader()
            writer.writerows(self.data)
            print(f'data has written to the {file_path}')

    def to_json(self, file_path = None, indent = None):
        if not self.data:
            raise ValueError('no data found')
        
        if file_path:
            with open(file_path, 'w') as file:
                json.dump(self.data, file, indent = indent)
            print(f'data has been written to the : {file_path}')

        else:
            return json.dumps(self.data, indent = indent)
        
    def read_json(self, file_path):
        if not os.path.exists(file_path):
            raise FileNotFoundError('could not find file')

        with open(file_path, 'r') as file:
            self.data = json.load(file)
        return self.data

    def to_excel(self, file_path):
        if not self.data:
            raise ValueError('no data found')
        
        work_book = Workbook()
        work_session = work_book.active

        columns = self.columns()
        work_session.append(columns)

        for row in self.data:
            work_session.append([row[col] for col in columns])
    
        work_book.save(file_path)
        print(f'data has been written to the : {file_path}')

    def read_excel(self, file_path):
        """Read data from an Excel file into the DataFrame."""

        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        columns = [cell.value for cell in sheet[1]]
        self.data = []

        for row in sheet.iter_rows(min_row=2, values_only=True):    
            if any(cell is not None for cell in row):
                row_dict = {columns[i]: cell for i, cell in enumerate(row)}
                self.data.append(row_dict)
        
        return self.data
    
    def dropna(self, axis=0, how='any'):
        if axis == 0:   
            if how == 'any':
                self.data = [row for row in self.data if not any(value is None or str(value).strip() == '' for value in row.values())]
            elif how == 'all':
                self.data = [row for row in self.data if not all(value is None or str(value).strip() == '' for value in row.values())]
        
        elif axis == 1:
            valid_columns = [col for col in self.columns() if any(row[col] is not None and str(row[col]).strip() != '' for row in self.data)]
            self.data = [{col: row[col] for col in valid_columns} for row in self.data]
        
        else:
            raise ValueError("axis must be 0 (rows) or 1 (columns)")
        
        return self

if __name__ == "__main__":
    data = ReadCsvJson('business.csv', dtype={'Identifier': float}, delimiter=',', skiprows=0, header=True)
    dropped_data = data.dropna(axis = 0, how = 'any')
    print(dropped_data)
    # print(dropeed_data[5 : 8])

    with open('dropna_testing.txt', 'w') as file:
        file.write(dropped_data.as_table())